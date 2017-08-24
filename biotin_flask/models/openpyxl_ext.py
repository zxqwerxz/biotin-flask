import csv
import os
from openpyxl import Workbook

from openpyxl.utils import column_index_from_string, coordinate_from_string
from openpyxl.styles import Border, Side


def delete_column(ws, delete_column):
    """
    This method empties a specified column from the worksheet. However,
    it still preserves the number of columns (ie. it doesn't actually
    delete any columns, but it empties one of the columns)

    This code is copy-pasted from https://stackoverflow.com/questions/22048016/delete-column-from-xls-file
    Author: John McGehee

    :param ws:                  (Object) Worksheet (not Workbook)
    :param delete_column:       (String) Column Letter OR (Integer) Column Number
    :return:                    (Object) Worksheet
    """
    if isinstance(delete_column, str):
        delete_column = column_index_from_string(delete_column)
    assert delete_column >= 1, 'Column numbers must be 1 or greater'

    for column in range(delete_column, ws.max_column + 1):
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=column).value = ws.cell(row=row, column=column+1).value

    return ws


def delete_row(ws, delete_row):
    """
    This method empties a specified row from the worksheet. However,
    it still preserves the number of row (ie. it doesn't actually
    delete any row, but it empties one of the rows)

    Copied and modified from delete_column

    :param ws:                  (Object) Worksheet (not Workbook)
    :param delete_column:       (Integer) Row Number
    :return:                    (Object) Worksheet
    """

    assert delete_row >= 1, 'Column numbers must be 1 or greater'

    for row in range(delete_row, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            ws.cell(row=row, column=column).value = ws.cell(row=row+1, column=column).value

    return ws


def wedge_column_right(ws, right, *wedge):
    """
    Move columns around in your worksheet!
    This method takes columns ("wedge") and moves them so that they
    are adjacent or to the right of another column ("right").
    The algorithm first shifts everything right of the column ("right") one
    column to the right. Then it assigns the values right of the
    column ("right") to the values in columns ("wedge"). Then it deletes
    the original columns ("wedge") using the method delete_column.
    Columns in wedge do not need to be adjacent to each other.
    The order that columns are provided in the args *wedge will
    be preserved when the columns are moved.

    :param ws:                  (Object) Worksheet (not Workbook)
    :param *wedge:              (String) Column Letter OR (Integer) Column Number
    :param right:               (String) Column Letter OR (Integer) Column Number
    :return:                    (Object) Worksheet
    """

    wedge_list = []
    for index,column in enumerate(wedge):
        if isinstance(column, str):
            wedge_list.append(column_index_from_string(column))
        assert wedge_list[index] >= 1, 'Column numbers must be 1 or greater'

    if isinstance(right, str):
        right = column_index_from_string(right)
    assert right >= 1, 'Column numbers must be 1 or greater'

    # If wedge is already to the right of right
    if wedge_list[0] == right + 1:
        return ws

    # Shift all columns to the right of right
    for column in range(ws.max_column,right,-1):
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=column+len(wedge_list)).value = ws.cell(row=row, column=column).value

    # Copy wedges to be adjacent to right
    for i, column in enumerate(wedge_list):
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=right+i+1).value = ws.cell(row=row, column=column+len(wedge_list)).value

    # Reverse sort wedge so that delete_column will be more efficient
    wedge_list = sorted(wedge_list, reverse=True)

    # Delete the original wedges
    for column in wedge_list:
        delete_column(ws, column+len(wedge_list))

    return ws


def csv_to_xlsx(csv_filepath):
    """
    This method converts csv files to .xlsx files because openpyxl can't
    deal with csv files.

    :param xls_wb:              (Path) path to csv file
    :return:                    None
    """

    xlsx = Workbook()
    xlsx_ws = xlsx.active

    # Copy elements in csv file to xlsx Workbook
    with open(csv_filepath, 'rb') as csvfile:
        csvreader = csv.reader(csvfile, delimiter="\t")
        for num, row in enumerate(csvreader):
            for col, element in enumerate(row):
                # Convert numbers into integers or floats
                if element.isdigit():
                    element = int(element)
                else:
                    try:
                        element = float(element)
                    except:
                        pass

                xlsx_ws.cell(row=num+1,column=col+1).value = element

    # Remove the old file
    os.remove(csv_filepath)

    # Save the new xlsx Workbook to the same path
    csv_filepath = csv_filepath + 'x'
    xlsx.save(csv_filepath)

def surround_border(ws, top_left, bottom_right):
    """
    This method surrounds the range with a thin border because openpyxl
    can't apply a style to a range of cells.
    This only works for a range of cells whose width is greater than 1
    and height is greater than 1 :(

    :param ws:                      (Object) Worksheet
    :param top_left:                (String) Cell location ie 'A1'
    :param bottom_right:            (String) Cell location ie 'A1'
    :return:                        (Object) Worksheet
    """

    top_left = list(coordinate_from_string(top_left))
    bottom_right = list(coordinate_from_string(bottom_right))
    top_left[0] = column_index_from_string(top_left[0])
    bottom_right[0] = column_index_from_string(bottom_right[0])

    for column in range(top_left[0], bottom_right[0] + 1):
        for row in range(top_left[1], bottom_right[1] + 1):
            if column == top_left[0]:
                ws.cell(row=row, column=column).border = Border(left=Side(border_style='thin'))
            if column == bottom_right[0]:
                ws.cell(row=row, column=column).border = Border(right=Side(border_style='thin'))
            if row == top_left[1]:
                ws.cell(row=row, column=column).border = Border(top=Side(border_style='thin'))
            if row == bottom_right[1]:
                ws.cell(row=row, column=column).border = Border(bottom=Side(border_style='thin'))

            # the four corners
            if column == top_left[0] and row == top_left[1]:
                ws.cell(row=row, column=column).border = Border(left=Side(border_style='thin'),
                                                                top=Side(border_style='thin'))
            if column == top_left[0] and row == bottom_right[1]:
                ws.cell(row=row, column=column).border = Border(left=Side(border_style='thin'),
                                                                bottom=Side(border_style='thin'))
            if column == bottom_right[0] and row == bottom_right[1]:
                ws.cell(row=row, column=column).border = Border(right=Side(border_style='thin'),
                                                                bottom=Side(border_style='thin'))
            if column == bottom_right[0] and row == top_left[1]:
                ws.cell(row=row, column=column).border = Border(right=Side(border_style='thin'),
                                                                top=Side(border_style='thin'))

    return ws