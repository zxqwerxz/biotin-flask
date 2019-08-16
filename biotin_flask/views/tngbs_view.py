"""Inputs tNGBS raw methylation data and final data format, Outputs the final data in the desired format

This module matches the CpG IDs in the tNGBS raw methylation data to the CpG IDs in the final excel sheet.
Output is the tNGBS data in the desired format.

Either a single excel file or multiple excel files can be provided.
"""

import os, io, csv, StringIO

from datetime import date
from flask import render_template, request, flash, send_file
from werkzeug import secure_filename
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from biotin_flask import app
from biotin_flask.models.openpyxl_ext import delete_column, wedge_column_right, csv_to_xlsx, delete_row, surround_border


@app.route('/misc/tngbs', methods=['GET', 'POST'])
def tngbs():
    """Handle GET or POST requests to the psq URL.

        Form args:
            f:      (File) list of .xlsx files uploaded by user
            ref:    (File) .xlsx file containing list of SNP IDs uploaded by user

        Returns:
            An excel file

        """

    # Render an empty form for GET request
    if request.method == 'GET':
        return render_template('tngbs/form.html')

    # Otherwise validate the form on a POST request and process uploaded files

    # Obtain files from the form
    f = request.files.getlist('raw')
    ref = request.files.getlist('format')

    # Throw error if one of the required files are missing
    if not f[0]:
        flash('A reqired field is missing', 'alert-warning')
        return render_template('tngbs/form.html')
    if not ref[0]:
        flash('A reqired field is missing', 'alert-warning')
        return render_template('tngbs/form.html')
    if len(ref) > 1:
        flash('Only one EpigenDX format file should be selected', 'alert-warning')
        return render_template('tngbs/form.html')

    # Validate the raw data files
    for file in f:
        filename = secure_filename(file.filename)
        filefront, extension = os.path.splitext(filename)
        if not extension == '.csv':
            flash('Only .csv files are allowed for the raw data files.', 'alert-warning')
            return render_template('tngbs/form.html')

    # Validate the reference file
    ref_filefront, ref_extension = os.path.splitext(secure_filename(ref[0].filename))
    if not ref_extension == '.xls' and not ref_extension == '.xlsx':
        flash('Only .xls and .xlsx files are allowed for the EpigenDx format file.', 'alert-warning')
        return render_template('tngbs/form.html')

    # Check if the reference file is saved on the server
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], 'tngbs_results.xlsx')
    if 'tngbs_results.xlsx' in os.listdir(app.config['UPLOAD_FOLDER']):
        os.remove(file_path)
        ref[0].save(file_path)
    else:
        ref[0].save(file_path)

    # Convert the reference file to xlsx so that openpyxl can work with it
    if ref_extension == '.xls':
        csv_to_xlsx(file_path)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], 'tngbs_results.xlsx')
    ref_book = load_workbook(file_path)

    # Open the ref file and store the FASTA IDs (e.g. TIGIT.2014) in a list
    ref_fasta_id = []
    for col in ref_book.active.iter_cols(min_col=5, max_col=ref_book.active.max_column, min_row=6, max_row=6):
        for cell in col:
            ref_fasta_id.append(cell.value)

    # Iterate through the raw data files
    fasta_id = []
    percent_meth = []
    coverage = []
    ref_percent_meth = []
    ref_coverage = []
    for file in f:
        try:
            stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
            reader = csv.reader(stream)
        except:
            flash('Unable to read csv file.', 'error')
            return render_template('tngbs/form.html')

        # Get first line, the list of FASTA IDs
        for value in next(reader)[1:]:
            fasta_id.append(value)

        # Iterate through the next rows
        is_meth = True
        for index, row in enumerate(reader):
            # If there's a line break, that means the data is switching from percent methylation values to read coverage values
            if row[0] == '':
                is_meth = False
                continue
            if is_meth:
                percent_meth[index] = []
                ref_percent_meth[index] = [None] * (len(ref_fasta_id)+1)
                ref_percent_meth[index][0] = row[0]
                for value in row:
                    percent_meth[index].append(value)
            if not is_meth:
                coverage[index] = []
                ref_coverage[index] = [None] * (len(ref_fasta_id)+1)
                ref_coverage[index][0] = row[0]
                for value in row[1:]:
                    coverage[index].append(value)

        # Match FASTA IDs
        """
        for index, value in enumerate(fasta_id):
            if value in ref_fasta_id:
                col_num = ref_fasta_id.index(value)
                for sample_id in ref_percent_meth:
                    ref_percent_meth[sample_id][col_num] = percent_meth[sample_id][index]
                for sample_id in ref_coverage:
                    ref_coverage[sample_id][col_num] = coverage[sample_id][index]
        """

    dest = StringIO.StringIO()
    writer = csv.writer(dest)

    return render_template('tngbs/form.html')

    """
    for sample_index in range(len(ref_percent_meth.keys()) + len(ref_coverage.keys()) + 1):
        min_row = 8
        min_col = 4
        for sample_id in ref_percent_meth:
            for item_index, item in enumerate(ref_percent_meth[sample_id]):
                ref_book.active.cell(row=sample_index+min_row, column=item_index+min_col).value = item
                print ref_book.active.cell(row=sample_index+min_row, column=item_index+min_col)
                print ref_book.active.cell(row=sample_index+min_row, column=item_index+min_col).value
        for sample_id in ref_coverage:
            for item_index, item in enumerate(ref_coverage[sample_id]):
                ref_book.active.cell(row=sample_index+min_row+1, column=item_index+min_col).value = item

    ref_book.save(file_path)

    return send_file(os.path.join(app.config['UPLOAD_FOLDER'], 'tngbs_results.xlsx'),
                     attachment_filename=ref_filefront+'_results.xlsx',
                     as_attachment=True)
    """

def delete_files(f):
    """

    :param f:       (File) list of files
    :return:        None
    """

    for file in f:
        filename = secure_filename(file.filename)
        filefront, extension = os.path.splitext(filename)
        extension = '.xlsx'
        filename = filefront + extension
        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], filename))