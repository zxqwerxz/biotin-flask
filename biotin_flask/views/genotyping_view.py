"""Inputs NGS raw genotyping data and outputs genotyping results

This module parses excel sheets uploaded by the user and reads variant call frequencies per sample.
It then outputs the genotyping results of each sample and each SNP

Either a single excel file or multiple excel files can be provided.
"""

import os, zipfile

from datetime import date
from flask import render_template, request, flash, send_file
from werkzeug import secure_filename
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from biotin_flask import app
from biotin_flask.models.openpyxl_ext import delete_column, wedge_column_right, csv_to_xlsx, delete_row, surround_border


@app.route('/misc/genotyping', methods=['GET', 'POST'])
def genotyping():
    """Handle GET or POST requests to the psq URL.

        Form args:
            f:      (File) list of .xlsx files uploaded by user
            ref:    (File) .xlsx file containing list of SNP IDs uploaded by user

        Returns:
            An excel file

        """
    # Check if there are any SNP ID lists in tmp/ref_snp_genotyping folder
    if len(os.listdir(os.path.join(app.config['UPLOAD_FOLDER'], 'ref_snp_genotyping'))) < 1:
        ref_exists = False
    else:
        ref_exists = True
    filenames = os.listdir(os.path.join(app.config['UPLOAD_FOLDER'], 'ref_snp_genotyping'))

    # Render an empty form for GET request
    if request.method == 'GET':
        return render_template('genotyping/form.html', ref_exists=ref_exists, filenames=filenames)

    # Otherwise validate the form on a POST request and process uploaded files

    # Obtain files from the form
    f = request.files.getlist('input')
    ref = request.files.getlist('reference')
    ref_sel = request.form.get('reference_sel')
    customer_id = request.form.get('customer_id')

    if not customer_id:
        customer_id = ''

    # Throw error if one of the required files are missing
    if not f[0]:
        flash('A reqired field is missing', 'error')
        return render_template('genotyping/form.html', ref_exists=ref_exists, filenames=filenames)
    if not ref[0] and ref_sel == 'None':
        flash('A reqired field is missing', 'error')
        return render_template('genotyping/form.html', ref_exists=ref_exists, filenames=filenames)

    # Validate and save variant call files
    for file in f:
        filename = secure_filename(file.filename)
        filefront, extension = os.path.splitext(filename)
        if not extension == '.xls':
            flash('Only .xls files are allowed.', 'error')
            return render_template('genotyping/form.html', ref_exists=ref_exists, filenames=filenames)
        path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
        file.save(path)

        # Convert the files in f to xlsx so that openpyxl can work with them
        csv_to_xlsx(path)
        extension = '.xlsx'
        file.filename = filefront + extension

    try:
        # Throw error if more than one snp list was provided
        if ref[0] and ref_sel != 'None':
            flash('Only one SNP ID reference file should be selected')
            delete_files(f)
            return render_template('genotyping/form.html', ref_exists=ref_exists, filenames=filenames)
        if len(ref) > 1:
            flash('Only one SNP ID reference file should be selected')
            delete_files(f)
            return render_template('genotyping/form.html', ref_exists=ref_exists, filenames=filenames)

        if ref_sel == 'None':
            # Validate the file object.
            filename = secure_filename(ref[0].filename)
            filefront, extension = os.path.splitext(filename)
            if not extension == '.xlsx':
                flash('Only .xlsx files are allowed.', 'error')
                delete_files(f)
                return render_template('genotyping/form.html', ref_exists=ref_exists, filenames=filenames)

            # Save the file
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ref_snp_genotyping', secure_filename(ref[0].filename))
            if secure_filename(ref[0].filename) in os.listdir(os.path.join(app.config['UPLOAD_FOLDER'], 'ref_snp_genotyping')):
                os.remove(file_path)
                ref[0].save(file_path)
                list_snp = load_workbook(file_path)
            else:
                ref[0].save(file_path)
                list_snp = load_workbook(file_path)
        else:
            list_snp = load_workbook(os.path.join(app.config['UPLOAD_FOLDER'], 'ref_snp_genotyping', ref_sel))

        # Open the ref file and store SNP IDs in a list
        snp_id = []
        for row in range(1, list_snp.active.max_row + 1):
            snp_id.append(list_snp.active['A{}'.format(row)].value)

        # Save the zip file
        try:
            if 'genotyping_results.zip' in os.listdir(app.config['UPLOAD_FOLDER']):
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'],'genotyping_results.zip'))
        except:
            delete_files(f)
            flash('Could not delete zip file on server', 'error')
            return render_template('genotyping/form.html', ref_exists=ref_exists, filenames=filenames)

        try:
            with zipfile.ZipFile(os.path.join(app.config['UPLOAD_FOLDER'], 'genotyping_results.zip'), 'w') as zipf:

                # Open the input files and begin main processing loop
                for file in f:
                    # Create the output excel file
                    output = Workbook()

                    file_wb = load_workbook(os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename)))
                    file_ws = file_wb.active

                    # Validate the spreadsheet
                    fields = ('Chrom', 'Position', 'Ref', 'Variant', 'Allele Call', 'Filter', 'Frequency', 'Quality', 'Filter',
                              'Type', 'Allele Source', 'Allele Name', 'Gene ID', 'Region Name', 'VCF Position',
                              'VCF Ref', 'VCF Variant', 'Original Coverage', 'Coverage', 'Filter', 'Coverage+',
                              'Filter', 'Coverage-', 'Filter', 'Allele Cov', 'Allele Cov+', 'Allele Cov-', 'Strand Bias',
                              'Filter', 'Common Signal Shift', 'Filter', 'Reference Signal Shift', 'Filter',
                              'Variant Signal Shift', 'Filter', 'Relative Read Quality', 'Filter', 'HP Length', 'Filter',
                              'Context Error+', 'Filter', 'Context Error-', 'Filter', 'Context Strand Bias', 'Filter',
                              'Sample Name', 'Barcode', 'Run Name', 'Allele', 'Location')
                    for column in range(1, 50):
                        if file_ws.cell(row=1, column=column).value != fields[column-1]:
                            delete_files(f)
                            flash('{} not formatted correctly'.format(secure_filename(file.filename)))
                            return render_template('genotyping/form.html', ref_exists=ref_exists, filenames=filenames)

                    # Rename some of the fields
                    file_ws['C1'].value = 'Ref Allele'
                    file_ws['D1'].value = 'Var Allele'
                    file_ws['G1'].value = 'Var Frequency'
                    file_ws['L1'].value = 'SNP ID'
                    file_ws['M1'].value = 'Gene'
                    file_ws['N1'].value = 'Assay ID'

                    # Remove the "comma glitch" in the SNP ID column
                    for row in range(2, file_ws.max_row + 1):
                        cell = file_ws['L{}'.format(row)].value
                        next_cell = file_ws['L{}'.format(row + 1)].value
                        if cell == '---' and next_cell[:2] == 'rs':
                            try:
                                file_ws['L{}'.format(row + 1)].value = next_cell[:next_cell.index(',')]
                            except:
                                pass

                    # Delete unnecessary columns
                    del_cols = ('AX','AW','AV','AS','AR','AQ','AP','AO','AN','AM','AL','AK','AJ','AI','AH','AG','AF','AE','AD',
                                'AC','X','V','T','R','Q','P','O','I','F')
                    for col in del_cols:
                        file_ws = delete_column(file_ws,col)

                    # Reorganize columns
                    file_ws = wedge_column_right(file_ws,'B','K','L','J')

                    # Iterate through the sheet rows
                    for row in range(2, file_ws.max_row + 1):
                        cell = file_ws['V{}'.format(row)] # genotyping result

                        # columns that represent various parameters
                        var_freq = file_ws['I{}'.format(row)].value
                        ref_allele = str(file_ws['F{}'.format(row)].value)
                        var_allele = str(file_ws['G{}'.format(row)].value)
                        coverage = file_ws['M{}'.format(row)].value

                        # Calculate the genotype
                        if var_freq < 10:
                            cell.value = ref_allele + '/' + ref_allele
                        elif var_freq > 90:
                            cell.value = var_allele + '/' + var_allele
                        else:
                            cell.value = ref_allele + '/' + var_allele
                        if coverage <= 5:
                            cell.value = 'NA'
                        elif coverage <= 30:
                            cell.value = 'MANUAL'

                        # Search SNP_ID in snp_id list and assign index to each row
                        index_cell = file_ws['W{}'.format(row)]
                        rs_id = str(file_ws['E{}'.format(row)].value)

                        try:
                            index_cell.value = snp_id.index(rs_id) + 1
                        except:
                            index_cell.value = 'na'

                    # Create new sheet for sorting
                    file_processed = file_wb.create_sheet('processed')

                    # Copy first row to the new sheet
                    for column in range(1, 23): # There are 22 columns, including the Genotyping column
                        file_processed.cell(row=1, column=column).value = file_ws.cell(row=1, column=column).value
                    file_processed['V1'].value = 'Genotype'
                    file_processed['W1'].value = 'Manual Call' # Empty column
                    file_processed['X1'].value = 'Average Coverage'


                    # Sort rows into the new processed sheet
                    sort_row = 1
                    total_coverage = 0
                    na_count = 0
                    for snp in range(1, len(snp_id) + 1):
                        for row in range(2, file_ws.max_row + 1):
                            if file_ws['W{}'.format(row)].value == snp:
                                sort_row += 1
                                # Copy row into new sheet
                                for column in range(1, 23): # This does not include the index column, column 23
                                    file_processed.cell(row=sort_row, column=column).value = file_ws.cell(row=row,
                                                                                                          column=column).value
                                # Test for triallele
                                if file_processed.cell(row=sort_row, column=5).value == file_processed.cell(row=sort_row-1,
                                                                                                            column=5).value:
                                    # Negate the previous count
                                    total_coverage -= file_processed.cell(row=sort_row-1, column=13).value

                                    tri_a = file_processed.cell(row=sort_row - 1, column=22).value
                                    tri_b = file_processed.cell(row=sort_row, column=22).value
                                    if tri_a == 'NA' or tri_b == 'NA':
                                        if '/' in tri_a:
                                            file_processed = delete_row(file_processed, sort_row)
                                            sort_row -= 1
                                        elif '/' in tri_b:
                                            file_processed = delete_row(file_processed, sort_row - 1)
                                            sort_row -=1
                                    elif tri_a == tri_b:
                                        file_processed = delete_row(file_processed, sort_row)
                                        sort_row -= 1
                                    elif '/' in tri_a and '/' in tri_b:
                                        split_a = tri_a.split('/')
                                        split_b = tri_b.split('/')
                                        if split_a[0] == split_b[0]:
                                            if split_a[1] == split_a[0]:
                                                file_processed = delete_row(file_processed, sort_row - 1)
                                                sort_row -= 1
                                            if split_b[1] == split_b[0]:
                                                file_processed = delete_row(file_processed, sort_row)
                                                sort_row -= 1
                                    else:
                                        # A human should figure this out
                                        file_processed.cell(row=sort_row - 1, column=22).value = 'TRIALLELE'
                                        file_processed.cell(row=sort_row, column=22).value = 'TRIALLELE'

                                total_coverage += file_processed.cell(row=sort_row, column=13).value

                    # Add an average coverage column
                    average_coverage = total_coverage/(sort_row - 1)
                    file_processed['X2'].value = average_coverage

                    # Make a thick border below the sort_row
                    sort_row += 1 # Skip a row
                    end_row = sort_row # Formatting will be added in output_sheet

                    # Copy the rest of the rows into the new processed sheet
                        # Rows labelled 'Hotspot'
                    for row in range(2, file_ws.max_row + 1):
                        if file_ws['W{}'.format(row)].value == 'na':
                            if file_ws['L{}'.format(row)].value == 'Hotspot':
                                sort_row += 1
                                # Copy row into new sheet
                                for column in range(1, 23):
                                    file_processed.cell(row=sort_row, column=column).value = file_ws.cell(row=row,
                                                                                                          column=column).value

                    # Copy the rest of the rows into the new processed sheet
                        # Rows labelled 'Novel'
                    for row in range(2, file_ws.max_row + 1):
                        if file_ws['W{}'.format(row)].value == 'na':
                            if file_ws['L{}'.format(row)].value == 'Novel':
                                sort_row += 1
                                # Copy row into new sheet
                                for column in range(1, 23):
                                    file_processed.cell(row=sort_row, column=column).value = file_ws.cell(row=row,
                                                                                                          column=column).value

                    # Create a new worksheet in output
                    output_sheet = output.create_sheet('Variant_Calls_' + file_processed['T2'].value)

                    # Because I'm really lazy, I'm adding this after the fact (This could be a lot more efficient)
                    del_cols = ['S','R','Q','P','O','N','L','K','J']
                    for col in del_cols:
                        delete_column(file_processed,col)

                    # Copy the processed worksheet into output
                    for row in range(1, file_processed.max_row + 1):
                        for column in range(1, 16): # There are 15 columns
                            output_sheet.cell(row=row, column=column).value = file_processed.cell(row=row, column=column).value
                            output_sheet.cell(row=row, column=column).alignment = Alignment(horizontal='center')

                    # Add the finishing formatting touches
                    output_sheet.cell(row=end_row, column=1).value = 'Extra SNPs'
                    output_sheet.cell(row=end_row, column=1).font = Font(bold=True)
                    for column in range(1, 16):  # There are 15 columns
                        output_sheet.cell(row=1, column=column).font = Font(bold=True)

                    # Create a second sheet
                    output_results = output.create_sheet('Results')
                    output_results['A1'] = 'These sequencing results were generated by EpigenDx, Inc.'
                    output_results['A1'].font = Font(bold=True)
                    output_results['A2'] = date.today().strftime('%m/%d/%y')
                    output_results['A3'] = 'Customer ID: ' + customer_id
                    output_results['A4'] = 'EpigenDx ID: ' + output_sheet['K2'].value
                    output_results['G6'] = 'Average Coverage = ' + str(average_coverage) + ' reads'
                    output_results['G8'] = 'Acceptance Criteria'
                    output_results['G8'].font = Font(bold=True)
                    output_results['G9'] = 'Homozygous with less than 20 reads = low quality'
                    output_results['G10'] = 'Heterozygous with less than 10 reads = low quality'
                    output_results['G11'] = 'Any SNP with less than 5 reads = NA'
                    output_results['G13'] = '# NA'
                    output_results['G14'] = 'Low Quality'

                    # Copy information from output_sheet to output_results
                    for row in range(1, end_row):
                        index = 1
                        for col in (1,2,3,5,13):
                            output_results.cell(row=row+4, column=index).value = output_sheet.cell(row=row, column=col).value
                            output_results.cell(row=row+4, column=index).alignment = Alignment(horizontal='center')
                            index += 1

                    # Bold
                    for column in range(1,6):
                        output_results.cell(row=5,column=column).font = Font(bold=True)

                    # Merging
                    for row in range(1,5):
                        output_results.merge_cells('A{}:E{}'.format(row,row))
                        output_results['A{}'.format(row)].alignment = Alignment(horizontal='center')
                    for row in range(8,12):
                        output_results.merge_cells('G{}:J{}'.format(row,row))
                        output_results['G{}'.format(row)].alignment = Alignment(horizontal='center')

                    # Borders
                    all_border = Border(left=Side(border_style='thin'),right=Side(border_style='thin'),
                                        top=Side(border_style='thin'),bottom=Side(border_style='thin'))
                    output_results = surround_border(output_results, 'A2', 'E4')
                    for column in range(1,6):
                        output_results.cell(row=1, column=column).border = all_border
                        output_results.cell(row=5, column=column).border = all_border
                    for column in range(7,11):
                        output_results.cell(row=8, column=column).border = all_border
                    output_results = surround_border(output_results,'A6','E'+str(output_results.max_row))
                    output_results = surround_border(output_results, 'G9', 'J11')

                    # Filling
                    greenFill = PatternFill(start_color='FFD8E4BC', end_color='FFD8E4BC',fill_type='solid')
                    grayFill = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
                    output_results['A1'].fill = greenFill
                    for row in range(2,5):
                        output_results['A{}'.format(row)].fill = grayFill
                    output_results['G8'].fill = grayFill

                    # Delete the first sheet of the output file
                    output.remove_sheet(output.active)

                    # Save the output file
                    filename = date.today().strftime('%m%d%y') + '_' + output_sheet['K2'].value
                    if customer_id == '':
                        filename += '_Results.xlsx'
                    else:
                        filename += '_' + customer_id + '_Results.xlsx'
                    out_path = os.path.join(app.config['UPLOAD_FOLDER'], 'genotyping_results', filename)
                    output.save(out_path)

                    # Write the file to a zip file
                    zipf.write(out_path, 'genotyping_results/' + os.path.basename(out_path))

                # Delete files from genotyping_results
                if not os.listdir(os.path.join(app.config['UPLOAD_FOLDER'], 'genotyping_results')):
                    delete_files(f)
                    flash('The server was not cleaned of generated results files', 'error')
                    return render_template('genotyping/form.html', ref_exists=ref_exists, filenames=filenames)
                for filename in os.listdir(os.path.join(app.config['UPLOAD_FOLDER'], 'genotyping_results')):
                    os.remove(os.path.join(app.config['UPLOAD_FOLDER'], 'genotyping_results', filename))
        except:
            delete_files(f)
            flash('An error happened while the zip file was open on the server', 'error')
            return render_template('genotyping/form.html', ref_exists=ref_exists, filenames=filenames)

        # Delete files from server
        delete_files(f)

        try:
            return send_file(os.path.join(app.config['UPLOAD_FOLDER'], 'genotyping_results.zip'),
                      attachment_filename='genotyping_results.zip',
                      as_attachment=True,
                      mimetype='zip')
        except:
            delete_files(f)
            flash('Could not send the zip file', 'error')
            return render_template('genotyping/form.html', ref_exists=ref_exists, filenames=filenames)

    except:
        # Catch all unknown exceptions
        delete_files(f)
        flash('An unknown error occurred :( Please let Eric know', 'error')
        return render_template('genotyping/form.html', ref_exists=ref_exists, filenames=filenames)


def delete_files(f):
    """

    :param f:       (File) list of files
    :return:        None
    """

    for file in f:
        filename = secure_filename(file.filename)
        filefront, extension = os.path.splitext(filename)
        extension = '.xlsx'
        filename = filefront + extension
        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], filename))
