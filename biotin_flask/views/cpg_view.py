#Add EPIC 450k array to all CpGs (excel file- Assay Validation)
#Lindsay Tomczak
#07-23-18

import xlrd, os, io, StringIO
from openpyxl import load_workbook
from flask import render_template, request, flash, make_response, send_file
from werkzeug import secure_filename
 
from biotin_flask import app
from biotin_flask.models.openpyxl_ext import csv_to_xlsx

__author__ = "Lindsay Tomczak"
__copyright__ = "Copyright (C) 2018, EpigenDx Inc."
__credits__ = ["Lindsay Tomczak"]
__version__ = "0.0.1"
__status__ = "Production"

@app.route("/misc/cgid", methods=["GET", "POST"])


#Main function that calls helper functions and gets data from html template page.
def cgid():
    """Handle GET or POST requests to the psq URL.

    Form args:
        f: (File) an excel file uploaded by user
        epicTab: (text) the name of the EPIC Array sheet in excel file
        hgVersion: (text) the name of the heading of the CG location
                    in the EPIC array sheet
        chrNum: (text) the number of the chromosome
        chrLocHeading: (text) the name of the location heading in All CpGs sheet 

        Returns:
        The same excel file with a new sheet added titled "New CpGs"

    """
    #Render an empty form for GET request.
    if request.method == 'GET':
        return render_template('cgid/form.html')

    #Otherwise validate the form on a POST request and process uploaded files.
    f= request.files.getlist('xlsx')
    for file in f:
        filename = secure_filename(file.filename)
        filefront, extension = os.path.splitext(filename)
        if not extension == '.xlsx' and not extension == '.xls':
            flash('Only .xlsx files are allowed.', 'alert-warning')
            return render_template('cgid/form.html')
        path= os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
        file.save(path)

    #Throw errors if file or required entries are missing or if more than one file
    #is entered.
    if not f[0]:
        flash('A required field is missing', 'alert-warning')
        return render_template('cgid/form.html')
    if len(f) > 1:
        flash('Only one file input is accepted', 'alert-warning')
        return render_template('cgid/form.html')

    epicTab= request.form.get('epic_name')
    if not epicTab[0]:
        flash('A required field is missing', 'alert-warning')
        return render_template('cgid/form.html')
    
    hgVersion= request.form.get('hg_version')
    if not hgVersion[0]:
        flash('A required field is missing', 'alert-warning')
        return render_template('cgid/form.html')

    chrNum= request.form.get('chr_num')
    if not chrNum[0]:
        flash('A required field is missing', 'alert-warning')
        return render_template('cgid/form.html')

    chrLocHeading= request.form.get('chr_heading')
    if not chrLocHeading[0]:
        flash('A required field is missing', 'alert-warning')
        return render_template('cgid/form.html')

    
    cSheet,eSheet=readFile(os.path.join(app.config['UPLOAD_FOLDER'],secure_filename(file.filename)),epicTab)
    eDict=makeEpicDict(eSheet,hgVersion,chrNum,chrLocHeading)
    a,b,c=getIndex(cSheet,chrLocHeading)
    cpgList=findCells(eDict,cSheet,a,b,c)
    newFile=writeNewSheet(os.path.join(app.config['UPLOAD_FOLDER'],secure_filename(file.filename)),cpgList)
    try:
        return send_file(newFile, as_attachment=True)
    except:
        flash('Could not send the file', 'alert-warning')
        return render_template('cgid/form.html')

    
#Begin Helper Functions
#Takes in the excel file name and the name of the epic 450k array tab
#since it is different on each excel (CpG tab usually stays the same),
#returns the desried sheets of the excel file, one for all CpGs and 450k.
def readFile(filename,epicName):
    excelBook= xlrd.open_workbook(filename, "rb")
    cpgSheet= excelBook.sheet_by_name("All CpGs")
    epicSheet= excelBook.sheet_by_name(epicName)
    return cpgSheet,epicSheet


#Takes in the EPIC 450k array sheet from the original excel file and
#the version of ensembl (hg38)/title of cgID location header and 
#goes through each row, creating a dictionary that maps the chromosome
#location to the cgID number, adding 'chr' with the number in front of the
#location is it is not there.
def makeEpicDict(epicSheet,version,chrNum,chrHeading):   
    epicDict={}
    rIndex,idIndex,hgIndex= getIndexEpic(epicSheet,version)
    for rowIndex in range(rIndex+1,epicSheet.nrows):
        keyValue= str(epicSheet.cell_value(rowIndex,hgIndex))
        keyValue= keyValue.lower()
        
        if "-" not in chrHeading:
            if "chr" not in keyValue:
                keyValueInt= int(epicSheet.cell_value(rowIndex,hgIndex))
                key= "chr" + str(chrNum) + ":" + str(keyValueInt)
            else:
                key = keyValue
            value= epicSheet.cell_value(rowIndex,idIndex)
            epicDict[key]=value
        
        if "-" in chrHeading:
            if "chr" not in keyValue:
                keyValueInt= int(epicSheet.cell_value(rowIndex,hgIndex))
                key= "chr" + str(chrNum) + ":" + str(keyValueInt+1)
            else:
                newCoord= getCoord(keyValue,chrNum)
                key = newCoord
            value= epicSheet.cell_value(rowIndex,idIndex)
            epicDict[key]=value
        rowIndex= rowIndex+1
   
    return(epicDict)


#Helper function for makeEpicDict, takes in the EPIC sheet from the 
#excel file and the version heading of that sheet containing the 
#coordinate values. Returns the row index for the labels,
#column index for the cg id column and the column index for the 
#coordinate column. 
def getIndexEpic(epicSheet,version):
    rowIndex=0
    idIndex=0
    hgIndex=0
    for j in range(epicSheet.nrows):
        for i in range(epicSheet.ncols):
            cellValue= epicSheet.cell_value(j,i)
            if cellValue == version:
                hgIndex=i
                rowIndex=j
            if cellValue == "IlmnID":
                idIndex=i
            else:
                i=i+1
        j=j+1
    return rowIndex,idIndex,hgIndex


#Helper function for makeEpicDict for the reverse (-) strand, takes in 
#the string of the cg coordinate from the EPIC file and the 
#chromosome number. Removes "Chr#:", adds 1 to the coordinate so it 
#matches the location in the CpG sheet. Returns the new coordinate.
def getCoord(cpgCoord,chrNum):
    chrBegin= "chr" + str(chrNum) + ":"
    coordNum= cpgCoord.replace(chrBegin,"")
    coordNum=int(coordNum) + 1
    return chrBegin + str(coordNum)
    


#Takes in the EPIC 450k array dictionary, the sheet in the workbook labelled
#"All CpGs", and the indicies from getIndex (for the correct columns and row
#to start with). Matches the chr location number in the dictionary to the
#one on the sheet and adds the cg# from the dictionary to the cpg# on 
#the sheet. If no cg#, just adds the cpg# to a list containing both the new
#modified cpg#s and the ones not modified and returns the list.
def findCells(epicDict,cpgSheet,rowI,cpgI,chrI):
    newCPGList=[]
    locKeys= epicDict.keys()
    for i in range((rowI +1),cpgSheet.nrows):
        cpgCell= cpgSheet.cell_value(i,cpgI)
        chrCell= cpgSheet.cell_value(i,chrI)
        chrCell=chrCell.lower()
        if chrCell in locKeys:
            cpgID= epicDict[chrCell]
            newCPGList.append(cpgCell + " " + cpgID)
        else:
            newCPGList.append(cpgCell)
        rowI=rowI + 1
    return newCPGList

 
#Takes in the excel workbook to be edited and a list of all of the new CpGs
#(orginal ones and ones changed by adding the EPIC cg #)
#creates a new sheet with the name "New CpGs" and pastes all CpGs in
#the firstcolumn.          
def writeNewSheet(filename,cpgList):
    excelFile= load_workbook(filename)
    newSheet=excelFile.create_sheet("New CpGs")
    for j in range(len(cpgList)):
        newSheet.append([cpgList[j] for i in range(1)])
    excelFile.save(filename)
    return filename



#Helper function gets the index for the row of labels in the cpg sheet,
#the column for CpGs and for the chr. location (heading needs to be given)
def getIndex(cpgSheet,chrHead):
    cpgIndex=0
    locIndex=0
    rowIndex=0
    for j in range(cpgSheet.nrows):
        for i in range(cpgSheet.ncols):
            cellValue= cpgSheet.cell_value(j,i)
            if cellValue == "CpG #":
                cpgIndex=i
                rowIndex=j
            if cellValue == chrHead:
                locIndex=i
            else:
                i=i+1
        j=j+1
    return(rowIndex,cpgIndex,locIndex)



